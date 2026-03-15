"""
Excel read/write for Hardball Dynasty amateur draft.
Maps between workbook sheets (Hitters, Pitchers) and draft order.
"""
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Lightweight Excel formula evaluator
# ---------------------------------------------------------------------------
# Handles the subset of Excel needed for typical Overall Projection formulas:
#   cell refs (A1, $A$1), ranges (H$1:P$1), SUMPRODUCT, SUM, ABS,
#   arithmetic (+, -, *, /, ^), parentheses, and numeric literals.
#   Formula cells are evaluated recursively (depth-limited).
# ---------------------------------------------------------------------------
_EVAL_MAX_DEPTH = 5


def _col_letter_to_num(col: str) -> int:
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _get_cell_as_float(ws, row: int, col: int, depth: int = 0) -> float:
    val = ws.cell(row, col).value
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("=") and depth < _EVAL_MAX_DEPTH:
        result = _eval_xl(ws, s[1:], depth + 1)
        return result if result is not None else 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _resolve_range(ws, range_str: str, depth: int = 0) -> list[float]:
    parts = range_str.strip().split(":")
    if len(parts) != 2:
        return []
    m1 = re.match(r"\$?([A-Z]+)\$?(\d+)", parts[0].strip())
    m2 = re.match(r"\$?([A-Z]+)\$?(\d+)", parts[1].strip())
    if not m1 or not m2:
        return []
    c1, r1 = _col_letter_to_num(m1.group(1)), int(m1.group(2))
    c2, r2 = _col_letter_to_num(m2.group(1)), int(m2.group(2))
    vals: list[float] = []
    if r1 == r2:
        for c in range(c1, c2 + 1):
            vals.append(_get_cell_as_float(ws, r1, c, depth))
    elif c1 == c2:
        for r in range(r1, r2 + 1):
            vals.append(_get_cell_as_float(ws, r, c1, depth))
    return vals


def _eval_xl(ws, expr: str, depth: int = 0) -> float | None:
    """Evaluate simplified Excel expression. Returns None on failure."""
    if depth > _EVAL_MAX_DEPTH:
        return None

    # SUMPRODUCT(range, range, ...)
    def _sp(m):
        args = [a.strip() for a in m.group(1).split(",")]
        ranges = [_resolve_range(ws, a, depth) for a in args]
        if not ranges or any(len(r) != len(ranges[0]) for r in ranges):
            return "0"
        total = 0.0
        for i in range(len(ranges[0])):
            p = 1.0
            for r in ranges:
                p *= r[i]
            total += p
        return str(total)

    expr = re.sub(r"SUMPRODUCT\(([^)]+)\)", _sp, expr, flags=re.IGNORECASE)

    # SUM(range_or_values)
    def _sm(m):
        inner = m.group(1).strip()
        parts = [p.strip() for p in inner.split(",")]
        total = 0.0
        for p in parts:
            if ":" in p:
                total += sum(_resolve_range(ws, p, depth))
            else:
                ref = re.match(r"\$?([A-Z]+)\$?(\d+)$", p)
                if ref:
                    total += _get_cell_as_float(
                        ws, int(ref.group(2)), _col_letter_to_num(ref.group(1)), depth
                    )
                else:
                    try:
                        total += float(p)
                    except ValueError:
                        pass
        return str(total)

    expr = re.sub(r"SUM\(([^)]+)\)", _sm, expr, flags=re.IGNORECASE)

    # ABS(expr)
    def _ab(m):
        inner = m.group(1).strip()
        try:
            return str(abs(float(inner)))
        except ValueError:
            return "0"

    expr = re.sub(r"ABS\(([^)]+)\)", _ab, expr, flags=re.IGNORECASE)

    # Replace remaining cell references with numeric values
    def _cr(m):
        col = m.group(1)
        row = int(m.group(2))
        return str(_get_cell_as_float(ws, row, _col_letter_to_num(col), depth))

    expr = re.sub(r"\$?([A-Z]+)\$?(\d+)", _cr, expr)

    expr = expr.replace("^", "**")

    try:
        return float(eval(expr))
    except Exception:
        return None


def _compute_projection(ws, data_row: int) -> float | None:
    """Evaluate the Overall Projection formula in column A for the given data row."""
    val = ws.cell(data_row, 1).value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s.startswith("="):
        try:
            return float(s)
        except ValueError:
            return None
    return _eval_xl(ws, s[1:])


# Sheet names and header row indices (1-based in Excel)
HITTERS_SHEET = "Hitters"
PITCHERS_SHEET = "Pitchers"
HITTERS_HEADER_ROW = 6   # row 6 has: total, Rnk, Player, Pos, B, T, Age, ...
PITCHERS_HEADER_ROW = 5  # row 5 has: Overall Projection, Rnk, Player, Pos, B, T, Age, ...

# Column names we care about (as they appear in header)
RNK = "Rnk"
PLAYER = "Player"
POS = "Pos"
B = "B"
T = "T"
AGE = "Age"
# Score columns used to rank when interleaving hitters and pitchers
HITTERS_SCORE_COL = "Overall Projection"
PITCHERS_SCORE_COL = "Overall Projection"

# Hitting block: columns B–P (basic info + 9 hitting ratings from "Projected Hitting Ratings" view).
# Positional keys (Rating_N) always exist in scraped rows; they're the primary lookup.
HITTERS_HITTING_LAYOUT: list[tuple[int, str, list[str]]] = [
    (2, "Rnk", ["Rating_1", "Rnk", "Rank"]),
    (3, "Player", ["Rating_2", "Player", "Player Name"]),
    (4, "Pos", ["Rating_3", "Pos", "Position"]),
    (5, "B", ["Rating_4", "B", "Bats"]),
    (6, "T", ["Rating_5", "T", "Throws"]),
    (7, "Age", ["Rating_6", "Age"]),
    (8, "Contact", ["Rating_7"]),
    (9, "Power", ["Rating_8"]),
    (10, "vs L", ["Rating_9"]),
    (11, "vs R", ["Rating_10"]),
    (12, "Batting Eye", ["Rating_11"]),
    (13, "Baserunning", ["Rating_12"]),
    (14, "Arm", ["Rating_13"]),
    (15, "Bunt", ["Rating_14"]),
    (16, "Overall", ["Rating_15"]),
]

# Fielding block starts at column Q (17).
# Direct copy-paste of the "Projected Fielding/General Ratings" view from the website.
FIELDING_BLOCK_START_COL = 17

# Pitchers sheet: single block starting at column B (2).
# 6 basic columns + 13 pitching ratings, all using positional keys from the scrape.
PITCHERS_LAYOUT: list[tuple[int, str, list[str]]] = [
    (2, "Rank", ["Rating_1", "Rnk", "Rank"]),
    (3, "Player", ["Rating_2", "Player", "Player Name"]),
    (4, "Position", ["Rating_3", "Pos", "Position"]),
    (5, "B", ["Rating_4", "B", "Bats"]),
    (6, "T", ["Rating_5", "T", "Throws"]),
    (7, "Age", ["Rating_6", "Age"]),
    (8, "Durability", ["Rating_7"]),
    (9, "Stamina", ["Rating_8"]),
    (10, "Control", ["Rating_9"]),
    (11, "vsL", ["Rating_10"]),
    (12, "vsR", ["Rating_11"]),
    (13, "Velocity", ["Rating_12"]),
    (14, "Groundball/Flyball Tendency", ["Rating_13"]),
    (15, "Pitch 1", ["Rating_14"]),
    (16, "Pitch 2", ["Rating_15"]),
    (17, "Pitch 3", ["Rating_16"]),
    (18, "Pitch 4", ["Rating_17"]),
    (19, "Pitch 5", ["Rating_18"]),
    (20, "Overall", ["Rating_19"]),
]

# Display headers for fielding columns, in website order (Fielding_1 = first column, etc.).
FIELDING_HEADERS: list[str] = [
    "Rank", "Player", "Pos", "B", "T", "Age",
    "Range", "Glove", "Arm Strength", "Arm Accuracy",
    "Pitch Calling", "Durability", "Health", "Speed",
    "Patience", "Temper", "Makeup", "Overall",
]


# ---------------------------------------------------------------------------
# Algorithm config: generate formulas from algorithm.json
# (polynomial coefficients + all individual/group weights)
# ---------------------------------------------------------------------------
ALGORITHM_FILE = Path(__file__).resolve().parent / "algorithm.json"
ALGORITHM_SHEET = "Algorithm"

# Rating name → Excel column letter for each sheet (fixed by template layout).
_HITTER_RATING_COL: dict[str, str] = {
    "Contact": "H", "Power": "I", "vs L": "J", "vs R": "K", "Batting Eye": "L",
    "Baserunning": "M", "Arm": "N", "Bunt": "O",
    "Range": "W", "Glove": "X", "Arm Strength": "Y", "Arm Accuracy": "Z",
    "Pitch Calling": "AA", "Durability": "AB", "Health": "AC",
    "Speed": "AD", "Patience": "AE", "Temper": "AF", "Makeup": "AG",
}
_PITCHER_RATING_COL: dict[str, str] = {
    "Durability": "H", "Stamina": "I", "Control": "J", "vsL": "K", "vsR": "L",
    "Velocity": "M", "GB/FB": "N",
    "Pitch 1": "O", "Pitch 2": "P", "Pitch 3": "Q", "Pitch 4": "R", "Pitch 5": "S",
}

_H_INTER_START = 35   # Column AI — first hitter intermediate column
_P_INTER_START = 21   # Column U  — first pitcher intermediate column
_H_REF_ROW = 5        # Hitter "perfect player" row (all 100s)
_P_REF_ROW = 4        # Pitcher "perfect player" row (all 100s)
_H_WEIGHT_ROW = 1     # Individual weight row for hitters
_H_CATCHER_ROW = 2    # Catcher-specific weight row for hitters
_H_GROUP_ROW = 3      # Group weight row for hitters
_P_WEIGHT_ROW = 1     # Individual weight row for pitchers
_P_GROUP_ROW = 2      # Group weight row for pitchers


def _load_algorithm_config() -> dict | None:
    """Load algorithm.json if it exists. Returns None if missing or invalid."""
    if not ALGORITHM_FILE.exists():
        return None
    try:
        import json
        with open(ALGORITHM_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception as exc:
        log.warning("Could not load algorithm.json: %s", exc)
        return None


def _col_letter_to_idx(col: str) -> int:
    """Convert Excel column letter(s) to 1-based index: A→1, Z→26, AA→27."""
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _fmtc(v: float) -> str:
    """Format a coefficient for Excel formulas (no scientific notation)."""
    if v == 0:
        return "0"
    s = f"{v:.12g}"
    if "e" in s or "E" in s:
        s = f"{v:.15f}".rstrip("0").rstrip(".")
    return s


def _poly(cell: str, coeff: dict) -> str:
    """Build the polynomial expression for one cell: a3*cell^3 + a2*cell^2 + a1*cell + a0."""
    a3, a2, a1, a0 = coeff.get("a3", 0), coeff.get("a2", 0), coeff.get("a1", 0), coeff.get("a0", 0)
    parts: list[str] = []
    if a3:
        c = _fmtc(a3)
        parts.append(f"({c})*{cell}^3" if a3 < 0 else f"{c}*{cell}^3")
    if a2:
        c = _fmtc(a2)
        parts.append(f"({c})*{cell}^2" if a2 < 0 else f"{c}*{cell}^2")
    if a1:
        c = _fmtc(a1)
        parts.append(f"({c})*{cell}" if a1 < 0 else f"{c}*{cell}")
    if a0:
        parts.append(_fmtc(a0))
    return "+".join(parts) if parts else "0"


def _weighted_poly(row: int, ratings: list[tuple[str, str]], coeff: dict) -> str:
    """SUM of weight_cell * poly(rating_cell) for a group.  ratings = [(col_letter, weight_cell), ...]"""
    return "+".join(f"{w}*({_poly(f'{c}{row}', coeff)})" for c, w in ratings)


def _simple_weighted(row: int, ratings: list[tuple[str, str]]) -> str:
    """(w1*r1 + w2*r2 + ...)/100 for groups that skip the polynomial."""
    inner = "+".join(f"{w}*{c}{row}" for c, w in ratings)
    return f"({inner})/100"


def _build_rating_refs(
    ratings_cfg: dict[str, float],
    col_map: dict[str, str],
    weight_row: int,
) -> list[tuple[str, str]]:
    """Convert {"Contact": 1.2, ...} → [("H", "H$1"), ...] for formula helpers."""
    refs: list[tuple[str, str]] = []
    for rname in ratings_cfg:
        col = col_map.get(rname)
        if col:
            refs.append((col, f"{col}${weight_row}"))
        else:
            log.warning("Algorithm config: unknown rating '%s'; skipping.", rname)
    return refs


def _write_weights_to_sheet(
    ws,
    groups_cfg: dict,
    col_map: dict[str, str],
    weight_row: int,
    catcher_row: int | None,
    group_row: int,
    inter_start: int,
) -> None:
    """Write individual rating weights (row 1), catcher weights (row 2), and
    group weights to their designated cells."""
    from openpyxl.utils import get_column_letter

    for gi, (gname, gdef) in enumerate(groups_cfg.items()):
        inter_col = inter_start + gi
        ws.cell(group_row, inter_col, float(gdef.get("group_weight", 1)))

        for rname, rweight in gdef.get("ratings", {}).items():
            col = col_map.get(rname)
            if col:
                ws.cell(weight_row, _col_letter_to_idx(col), float(rweight))

        if catcher_row:
            for rname, rweight in gdef.get("catcher_ratings", {}).items():
                col = col_map.get(rname)
                if col:
                    ws.cell(catcher_row, _col_letter_to_idx(col), float(rweight))


def _generate_group_formula(
    row: int,
    gdef: dict,
    col_map: dict[str, str],
    weight_row: int,
    catcher_row: int | None,
    coeff: dict,
    is_ref: bool = False,
) -> str:
    """Generate the intermediate-column formula for one group at one row."""
    method = gdef.get("method", "polynomial")
    refs = _build_rating_refs(gdef.get("ratings", {}), col_map, weight_row)

    if method == "polynomial":
        expr = _weighted_poly(row, refs, coeff)
    else:
        expr = _simple_weighted(row, refs)

    catcher_cfg = gdef.get("catcher_ratings")
    if catcher_cfg and catcher_row:
        catcher_refs = _build_rating_refs(catcher_cfg, col_map, catcher_row)
        if method == "polynomial":
            catcher_expr = _weighted_poly(row, catcher_refs, coeff)
        else:
            catcher_expr = _simple_weighted(row, catcher_refs)
        cond_col = col_map.get(gdef.get("catcher_condition", "Pitch Calling"), "AA")
        threshold = int(gdef.get("catcher_threshold", 50))
        cond_op = ">" if is_ref else "<"
        return f"=IF({cond_col}{row}{cond_op}{threshold},{expr},{catcher_expr})"

    return f"={expr}"


def _generate_col_a(
    row: int,
    n_groups: int,
    inter_start: int,
    group_row: int,
    ref_row: int,
) -> str:
    """Column A = (SUM of intermediate*group_weight)*100 / (SUM of ref*group_weight)."""
    from openpyxl.utils import get_column_letter
    parts_num: list[str] = []
    parts_den: list[str] = []
    for gi in range(n_groups):
        ic = get_column_letter(inter_start + gi)
        gw = f"{ic}${group_row}"
        parts_num.append(f"{ic}{row}*{gw}")
        parts_den.append(f"{ic}${ref_row}*{gw}")
    return f"=({'+'.join(parts_num)})*100/({'+'.join(parts_den)})"


def _apply_algorithm_formulas(
    wb: openpyxl.Workbook,
    algo: dict,
    n_hitters: int,
    n_pitchers: int,
) -> None:
    """Write weights, intermediate-column formulas, and column-A formulas
    from algorithm.json (polynomial + all weights)."""
    coeff = algo.get("polynomial", {})
    if not coeff:
        return

    for sheet_name, section, col_map, n_rows, header_row, ref_row, \
            inter_start, w_row, c_row, g_row in [
        (HITTERS_SHEET, "hitters", _HITTER_RATING_COL, n_hitters,
         HITTERS_HEADER_ROW, _H_REF_ROW, _H_INTER_START,
         _H_WEIGHT_ROW, _H_CATCHER_ROW, _H_GROUP_ROW),
        (PITCHERS_SHEET, "pitchers", _PITCHER_RATING_COL, n_pitchers,
         PITCHERS_HEADER_ROW, _P_REF_ROW, _P_INTER_START,
         _P_WEIGHT_ROW, None, _P_GROUP_ROW),
    ]:
        groups_cfg = algo.get(section, {}).get("groups", {})
        if not groups_cfg or sheet_name not in wb.sheetnames or n_rows == 0:
            continue

        ws = wb[sheet_name]
        n_groups = len(groups_cfg)

        _write_weights_to_sheet(ws, groups_cfg, col_map, w_row, c_row, g_row, inter_start)

        all_rows = [ref_row] + [header_row + 1 + i for i in range(n_rows)]
        for r in all_rows:
            is_ref = (r == ref_row)
            for gi, (gname, gdef) in enumerate(groups_cfg.items()):
                fml = _generate_group_formula(r, gdef, col_map, w_row, c_row, coeff, is_ref)
                ws.cell(r, inter_start + gi, fml)
            if r != ref_row:
                ws.cell(r, 1, _generate_col_a(r, n_groups, inter_start, g_row, ref_row))

        # Write group labels on the header row at the intermediate columns
        from openpyxl.utils import get_column_letter
        for gi, gname in enumerate(groups_cfg):
            ws.cell(header_row, inter_start + gi, gname)

    # ── Write coefficients to Algorithm tab for visibility ──
    if ALGORITHM_SHEET not in wb.sheetnames:
        wb.create_sheet(ALGORITHM_SHEET)
    ws_a = wb[ALGORITHM_SHEET]
    ws_a.cell(1, 13, "a3")
    ws_a.cell(1, 14, "a2")
    ws_a.cell(1, 15, "a1")
    ws_a.cell(1, 16, "a0")
    ws_a.cell(2, 13, coeff.get("a3", 0))
    ws_a.cell(2, 14, coeff.get("a2", 0))
    ws_a.cell(2, 15, coeff.get("a1", 0))
    ws_a.cell(2, 16, coeff.get("a0", 0))

    log.info("Formulas and weights written from algorithm.json.")


def _row_value_for_keys(row: dict[str, Any], keys: list[str]) -> Any:
    """Return row[key] for the first key that exists (case-insensitive key match)."""
    row_lower = {str(k).strip().lower(): v for k, v in row.items() if k is not None and str(k).strip()}
    for k in keys:
        if k is None or not str(k).strip():
            continue
        if k in row:
            return row[k]
        kl = str(k).strip().lower()
        for rk, rv in row_lower.items():
            if rk == kl:
                return rv
    return None


def _header_to_col(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, name: str) -> int | None:
    """Return 1-based column index for header name (exact match), or None."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip() == name:
            return col
    return None


def _header_to_col_ic(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, name: str) -> int | None:
    """Return 1-based column index for header name (case-insensitive), or None."""
    name_lower = name.lower()
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is not None and str(val).strip().lower() == name_lower:
            return col
    return None


def _parse_score(val: Any) -> float:
    """Parse a cell value as a numeric score; return 0 if missing or invalid."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def validate_template(path: str | Path) -> list[str]:
    """
    Check that the workbook has the expected structure. Returns list of error messages (empty if OK).
    """
    path = Path(path)
    errors: list[str] = []
    if not path.exists():
        return [f"File not found: {path}"]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [f"Cannot open workbook: {e}"]
    for sheet_name, header_row, score_col in [
        (HITTERS_SHEET, HITTERS_HEADER_ROW, HITTERS_SCORE_COL),
        (PITCHERS_SHEET, PITCHERS_HEADER_ROW, PITCHERS_SCORE_COL),
    ]:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        if _header_to_col(ws, header_row, PLAYER) is None and _header_to_col_ic(ws, header_row, PLAYER) is None:
            errors.append(f"Sheet '{sheet_name}' has no 'Player' column in header row {header_row}")
        if _header_to_col(ws, header_row, score_col) is None and _header_to_col_ic(ws, header_row, score_col) is None:
            errors.append(f"Sheet '{sheet_name}' has no '{score_col}' column in header row {header_row}")
    wb.close()
    return errors


def get_draft_order_from_excel(path: str | Path) -> list[str]:
    """
    Read draft order from the Master List sheet (already sorted by Adjusted Score).
    Falls back to Hitters + Pitchers sheets sorted by Overall Projection if no Master List.
    Returns list of player names in draft order.
    """
    path = Path(path)
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Prefer the Master List — it's already sorted by adjusted score
    if MASTER_LIST_SHEET in wb.sheetnames:
        ws = wb[MASTER_LIST_SHEET]
        player_col = _header_to_col(ws, 1, "Player") or _header_to_col_ic(ws, 1, "Player")
        if player_col:
            names: list[str] = []
            for row in range(2, ws.max_row + 1):
                name = ws.cell(row, player_col).value
                if name is None or not str(name).strip():
                    continue
                names.append(str(name).strip())
            if names:
                wb.close()
                return names

    # Fallback: read from Hitters + Pitchers, sort by Overall Projection
    scored: list[tuple[float, str]] = []
    for sheet_name, header_row, score_col in [
        (HITTERS_SHEET, HITTERS_HEADER_ROW, HITTERS_SCORE_COL),
        (PITCHERS_SHEET, PITCHERS_HEADER_ROW, PITCHERS_SCORE_COL),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        player_col = _header_to_col(ws, header_row, PLAYER)
        score_col_idx = _header_to_col(ws, header_row, score_col) or _header_to_col_ic(ws, header_row, score_col)
        if player_col is None:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row, player_col).value
            if name is None or not str(name).strip():
                continue
            score = _parse_score(ws.cell(row, score_col_idx).value) if score_col_idx else 0.0
            scored.append((score, str(name).strip()))

    wb.close()
    scored.sort(key=lambda x: (-x[0], x[1]))
    return [name for _, name in scored]


def _normalize_name(name: str) -> str:
    """Normalize for matching (strip, single spaces)."""
    return " ".join(str(name).strip().split())


def get_excel_row_ranges(path: str | Path) -> tuple[tuple[int, int], tuple[int, int]]:
    """Return (hitters_data_start, hitters_data_end), (pitchers_data_start, pitchers_data_end) (1-based)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hitters = (HITTERS_HEADER_ROW + 1, 0)
    pitchers = (PITCHERS_HEADER_ROW + 1, 0)
    if HITTERS_SHEET in wb.sheetnames:
        ws = wb[HITTERS_SHEET]
        hitters = (HITTERS_HEADER_ROW + 1, ws.max_row)
    if PITCHERS_SHEET in wb.sheetnames:
        ws = wb[PITCHERS_SHEET]
        pitchers = (PITCHERS_HEADER_ROW + 1, ws.max_row)
    wb.close()
    return hitters, pitchers


def _cols_with_error_or_empty(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> list[int]:
    """Return 1-based column indices where header is #VALUE!, empty, or looks like an error (for overwriting)."""
    result = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is None:
            result.append(col)
            continue
        s = str(val).strip()
        if s == "" or s == "#VALUE!" or "VALUE" in s.upper() or "REF" in s.upper():
            result.append(col)
    return result


def _write_hitters_sheet_fixed(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    rows: list[dict[str, Any]],
) -> None:
    """
    Write hitters data using the fixed template layout.
    Hitting block (B–P): basic info + 9 hitting ratings.
    Fielding block (Q onwards): direct copy-paste of the Fielding/General view (all columns in website order).
    """
    # Hitting block
    for col_idx, header_label, keys in HITTERS_HITTING_LAYOUT:
        ws.cell(header_row, col_idx, header_label)
        for i, row in enumerate(rows):
            val = _row_value_for_keys(row, keys)
            if val is not None:
                ws.cell(header_row + 1 + i, col_idx, val)

    # Fielding block: all Fielding_N keys in order, starting at column Q
    if not rows:
        return
    fielding_keys = sorted(
        [k for k in rows[0] if k.startswith("Fielding_")],
        key=lambda k: int(k.split("_")[1]),
    )
    if not fielding_keys:
        return
    start_col = FIELDING_BLOCK_START_COL
    for offset, fkey in enumerate(fielding_keys):
        col = start_col + offset
        header = FIELDING_HEADERS[offset] if offset < len(FIELDING_HEADERS) else fkey
        ws.cell(header_row, col, header)
        for i, row in enumerate(rows):
            val = row.get(fkey)
            if val is not None:
                ws.cell(header_row + 1 + i, col, val)


MASTER_LIST_SHEET = "Master List"

# Scouting trust factor parameters.
# trust = MIN_TRUST + (1 - MIN_TRUST) * (budget / max_budget) ^ CURVE
_SCOUTING_MAX_BUDGET = 20.0  # HBD's fixed max scouting budget ($M)


def _raw_scouting_trust(budget: float, min_trust: float, curve: float) -> float:
    """Raw trust value before normalization. Used internally."""
    ratio = max(0.0, min(1.0, budget / _SCOUTING_MAX_BUDGET))
    if ratio == 0:
        return min_trust
    return min_trust + (1 - min_trust) * (ratio ** curve)


def _classify_player(age: Any, player_class: str = "") -> str:
    """
    Classify a player as 'college' or 'high_school'.
    Uses Class field first (FR/SO/JR/SR = college, -- = HS), then falls back to age (19+ = college).
    """
    pc = str(player_class or "").strip().upper()
    if pc in ("FR", "SO", "JR", "SR"):
        return "college"
    if pc == "--":
        return "high_school"
    try:
        a = int(age)
    except (TypeError, ValueError):
        return "high_school"
    if a >= 19:
        return "college"
    return "high_school"


def _signability_factor(text: str, raw_overall: float = 0, cfg: dict[str, float] | None = None) -> float:
    """
    Return a multiplier (0–1) based on signability text from the Background Info view.
    "First round" / "first five rounds" penalties only apply if the player isn't good enough
    to actually go in that range (i.e. they'd leave for college instead of signing).
    "Probably won't sign" and "Unknown" are near-zero to avoid drafting them.
    All penalty values are configurable via credentials.env (SIGN_* keys).
    """
    if cfg is None:
        from credentials import get_signability_config
        cfg = get_signability_config()

    t = (text or "").strip().lower()
    if not t:
        return 1.0
    if "will sign for slot" in t or "looking to sign" in t:
        return cfg["will_sign"]
    if "drafted in the first round" in t:
        return 1.0 if raw_overall >= cfg["first_round_threshold"] else cfg["first_round"]
    if "drafted in the first five" in t:
        return 1.0 if raw_overall >= cfg["first_five_threshold"] else cfg["first_five"]
    if "may sign if the deal is right" in t:
        return cfg["may_sign"]
    if "undecided" in t:
        return cfg["undecided"]
    if "probably won't sign" in t:
        return cfg["probably_wont"]
    if "unknown" in t or "wasn't scouted" in t:
        return cfg["unknown"]
    return cfg["fallback"]


BACKGROUND_SHEET = "Background Info"
BACKGROUND_HEADERS = ["Rnk", "Player", "Pos", "B", "T", "Age", "Hometown", "School", "Class", "Signability"]


def _write_background_sheet(
    wb: openpyxl.Workbook,
    background_rows: list[dict[str, Any]],
) -> None:
    """Write a Background Info tab with all players' background data."""
    if BACKGROUND_SHEET in wb.sheetnames:
        del wb[BACKGROUND_SHEET]
    ws = wb.create_sheet(BACKGROUND_SHEET)
    for col_idx, header in enumerate(BACKGROUND_HEADERS, start=1):
        ws.cell(1, col_idx, header)
    for i, row in enumerate(background_rows):
        for col_idx, header in enumerate(BACKGROUND_HEADERS, start=1):
            val = row.get(header)
            if val is not None:
                ws.cell(i + 2, col_idx, val)


def _sort_master_list_via_excel(path: Path) -> bool:
    """
    Open the saved workbook in Excel via COM, recalculate all formulas,
    sort the Master List by Adjusted Score (column A) descending,
    delete rows where the Adjusted Score is an error (#VALUE!), and save.
    Returns True on success.
    """
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        log.info("pywin32 not installed; Master List sort requires: pip install pywin32")
        return False

    excel = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        abs_path = str(path.resolve())
        wb = excel.Workbooks.Open(abs_path)
        excel.CalculateFull()

        ws = wb.Sheets(MASTER_LIST_SHEET)
        last_row = ws.Cells(ws.Rows.Count, 6).End(-4162).Row  # xlUp; col F = Player name

        if last_row > 1:
            # Sort by Adjusted Score descending (errors sink to the bottom)
            sort_range = ws.Range(f"A1:J{last_row}")
            sort_range.Sort(
                Key1=ws.Range("A2"),
                Order1=2,       # xlDescending
                Header=1,       # xlYes
                OrderCustom=1,
                MatchCase=False,
                Orientation=1,  # xlTopToBottom
            )

            # Delete rows from the bottom up where Adjusted Score is an error or non-numeric
            deleted = 0
            for r in range(last_row, 1, -1):
                val = ws.Cells(r, 1).Value
                if val is None or not isinstance(val, (int, float)) or val <= 0:
                    ws.Rows(r).Delete()
                    deleted += 1

            log.info("Excel COM: sorted Master List (%d rows, removed %d error rows).",
                     last_row - 1 - deleted, deleted)

        wb.Save()
        wb.Close()
        excel.Quit()
        pythoncom.CoUninitialize()
        return True

    except Exception as e:
        log.warning("Excel COM sort failed: %s", e)
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        return False


def _write_master_list(
    wb: openpyxl.Workbook,
    hitters_rows: list[dict[str, Any]],
    pitchers_rows: list[dict[str, Any]],
) -> None:
    """
    Create a Master List tab with all players (hitters + pitchers).
    Adjusted Score (col A) is an Excel formula: Overall Projection × Trust × Signability.
    Initial row order uses raw Overall as a rough approximation; the caller should
    follow up with _sort_master_list_via_excel() to get the real sort by formula values.
    """
    from credentials import get_scouting_config, get_signability_config

    scouting = get_scouting_config()
    sign_cfg = get_signability_config()
    min_trust = scouting["min_trust"]
    curve = scouting["curve"]

    raw_trusts = {
        "college": _raw_scouting_trust(scouting["college"], min_trust, curve),
        "high_school": _raw_scouting_trust(scouting["high_school"], min_trust, curve),
    }
    best = max(raw_trusts.values())
    trust_factors = {k: v / best for k, v in raw_trusts.items()}

    if MASTER_LIST_SHEET in wb.sheetnames:
        del wb[MASTER_LIST_SHEET]
    ws = wb.create_sheet(MASTER_LIST_SHEET)

    players: list[dict[str, Any]] = []

    for i, row in enumerate(hitters_rows):
        src_row = HITTERS_HEADER_ROW + 1 + i
        name = str(_row_value_for_keys(row, ["Rating_2", "Player", "Player Name"]) or "")
        pos = str(_row_value_for_keys(row, ["Rating_3", "Pos", "Position"]) or "")
        age = _row_value_for_keys(row, ["Rating_6", "Age"])
        player_class = str(row.get("Class", "") or "")
        signability = str(row.get("Signability", "") or "")
        raw = _parse_score(_row_value_for_keys(row, ["Rating_15"]) or 0)
        category = _classify_player(age, player_class)
        scout_trust = trust_factors[category]
        sign_factor = _signability_factor(signability, raw, sign_cfg)
        players.append({
            "sort_key": raw * scout_trust * sign_factor,
            "raw": raw, "scout": scout_trust, "sign": sign_factor,
            "name": name, "pos": pos, "type": "Hitter", "cat": category,
            "sig": signability, "sheet": HITTERS_SHEET, "src_row": src_row,
        })

    for i, row in enumerate(pitchers_rows):
        src_row = PITCHERS_HEADER_ROW + 1 + i
        name = str(_row_value_for_keys(row, ["Rating_2", "Player", "Player Name"]) or "")
        pos = str(_row_value_for_keys(row, ["Rating_3", "Pos", "Position"]) or "")
        age = _row_value_for_keys(row, ["Rating_6", "Age"])
        player_class = str(row.get("Class", "") or "")
        signability = str(row.get("Signability", "") or "")
        raw = _parse_score(_row_value_for_keys(row, ["Rating_19"]) or 0)
        category = _classify_player(age, player_class)
        scout_trust = trust_factors[category]
        sign_factor = _signability_factor(signability, raw, sign_cfg)
        players.append({
            "sort_key": raw * scout_trust * sign_factor,
            "raw": raw, "scout": scout_trust, "sign": sign_factor,
            "name": name, "pos": pos, "type": "Pitcher", "cat": category,
            "sig": signability, "sheet": PITCHERS_SHEET, "src_row": src_row,
        })

    players.sort(key=lambda p: (-p["sort_key"], p["name"]))

    # Column layout:
    #   A = Adjusted Score  (formula: B × D × E — always correct when opened in Excel)
    #   B = Overall Projection  (formula referencing source sheet column A)
    #   C = Raw Overall  (HBD's raw rating number)
    #   D = Scouting Trust  (multiplier from scouting budget config)
    #   E = Signability Factor  (multiplier from signability text)
    #   F = Player
    #   G = Pos
    #   H = Type  (Hitter / Pitcher)
    #   I = Category  (college / high_school)
    #   J = Signability  (raw text from Background Info)
    headers = [
        "Adjusted Score", "Overall Projection", "Raw Overall",
        "Scouting Trust", "Signability Factor",
        "Player", "Pos", "Type", "Category", "Signability",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(1, col_idx, h)

    for i, p in enumerate(players):
        r = i + 2
        ws.cell(r, 1, f"=B{r}*D{r}*E{r}")
        ws.cell(r, 2, f"='{p['sheet']}'!A{p['src_row']}")
        ws.cell(r, 3, p["raw"])
        ws.cell(r, 4, round(p["scout"], 3))
        ws.cell(r, 5, round(p["sign"], 2))
        ws.cell(r, 6, p["name"])
        ws.cell(r, 7, p["pos"])
        ws.cell(r, 8, p["type"])
        ws.cell(r, 9, p["cat"])
        ws.cell(r, 10, p["sig"])


def write_draft_data_to_excel(
    path: str | Path,
    hitters_rows: list[dict[str, Any]],
    pitchers_rows: list[dict[str, Any]],
    background_rows: list[dict[str, Any]] | None = None,
    output_path: str | Path | None = None,
) -> None:
    """
    Write scraped draft pool data into the Excel file.

    Flow:
      1. Write Hitters, Pitchers, Background Info, and Master List sheets via openpyxl.
         The Master List's Adjusted Score is a formula (=B*D*E) so it can only be
         evaluated by Excel; initial row order is an approximation (raw Overall).
      2. Save the workbook.
      3. Open the saved file in Excel via COM, recalculate all formulas, sort the
         Master List by Adjusted Score descending, and save.  Players whose
         Overall Projection is a #VALUE! error end up at the bottom automatically.

    If Excel COM is unavailable the file is still valid — formulas compute when
    the user opens it in Excel, and they can sort manually.
    """
    path = Path(path)
    save_to = Path(output_path) if output_path else path
    save_to.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(path)

    if HITTERS_SHEET in wb.sheetnames and hitters_rows:
        _write_hitters_sheet_fixed(wb[HITTERS_SHEET], HITTERS_HEADER_ROW, hitters_rows)

    if PITCHERS_SHEET in wb.sheetnames:
        ws = wb[PITCHERS_SHEET]
        for col_idx, header_label, keys in PITCHERS_LAYOUT:
            ws.cell(PITCHERS_HEADER_ROW, col_idx, header_label)
            for i, row in enumerate(pitchers_rows):
                val = _row_value_for_keys(row, keys)
                if val is not None:
                    ws.cell(PITCHERS_HEADER_ROW + 1 + i, col_idx, val)

    # Generate column-A formulas from algorithm.json (if present)
    algo = _load_algorithm_config()
    if algo:
        _apply_algorithm_formulas(wb, algo, len(hitters_rows), len(pitchers_rows))

    # Wrap column-A formulas with IFERROR so errors show 0 instead of #VALUE!
    for sheet_name, header_row, n_rows in [
        (HITTERS_SHEET, HITTERS_HEADER_ROW, len(hitters_rows)),
        (PITCHERS_SHEET, PITCHERS_HEADER_ROW, len(pitchers_rows)),
    ]:
        if sheet_name in wb.sheetnames and n_rows:
            ws = wb[sheet_name]
            for r in range(header_row + 1, header_row + 1 + n_rows):
                cell = ws.cell(r, 1)
                v = cell.value
                if v and str(v).startswith("=") and not str(v).upper().startswith("=IFERROR"):
                    cell.value = f"=IFERROR({str(v)[1:]},0)"

    if background_rows:
        _write_background_sheet(wb, background_rows)

    _write_master_list(wb, hitters_rows, pitchers_rows)

    wb.save(save_to)
    wb.close()
    log.info("Saved workbook: %s", save_to)

    # Let Excel recalculate formulas and sort the Master List by Adjusted Score.
    if _sort_master_list_via_excel(save_to):
        log.info("Master List sorted by Adjusted Score via Excel.")
    else:
        log.warning("Excel COM sort unavailable; open the file in Excel and sort "
                     "Master List by column A descending.")


def append_draft_order_sheet(path: str | Path, order: list[str], sheet_name: str = "DraftOrder") -> None:
    """Append a simple 'DraftOrder' sheet with one column of player names in order."""
    path = Path(path)
    wb = openpyxl.load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, "Order")
    ws.cell(2, 1, "Player")
    for i, name in enumerate(order, start=1):
        ws.cell(i + 2, 1, name)
    wb.save(path)
    wb.close()
